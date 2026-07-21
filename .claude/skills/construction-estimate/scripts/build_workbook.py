#!/usr/bin/env python3
"""Build the estimate workbook from a takeoff JSON file.

Produces the internal working estimate (every line, rates, hours, sources) and
the client rollup (division totals only). The client sheet deliberately carries
no rates, hours or margin — see CLAUDE.md.

    build_workbook.py takeoff.json -o estimate.xlsx

Takeoff JSON:

    {
      "project":     "2212 E 7th Ave",
      "client":      "Client Name",
      "prepared_by": "Caisley Developments",
      "date":        "2026-07-21",
      "drawings":    "VFA, Building Permit set, issued 2024-09-06",
      "drawings_current": false,
      "blended_rate": 68.21,
      "contingency": 0.05,
      "fee":         0.10,
      "gst":         0.05,
      "lines": [
        {"code": "2.2", "desc": "Remove existing drywall", "qty": 2140,
         "unit": "sq.ft.", "rate": 7.00, "hours": 2.14,
         "kind": "self", "source": "D200"}
      ],
      "assumptions": ["..."],
      "exclusions":  ["..."],
      "questions":   [{"q": "...", "assumed": "...", "affects": "9.4"}],
      "risks":       [{"item": "...", "why": "..."}]
    }

`kind` is one of self | subtrade | quote | pc | excluded. An excluded line
carries text instead of a cost ("NOT INCLUDED", "BY OWNER", "BY OTHERS") and
contributes nothing to the total, but stays visible.
"""

import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl not installed. Run: pip3 install openpyxl")

DIVISIONS = {
    0: "BIDDING AND CONTRACT REQUIREMENTS",
    1: "GENERAL REQUIREMENTS",
    2: "SITEWORK",
    3: "CONCRETE",
    4: "MASONRY",
    5: "METALS",
    6: "WOOD AND PLASTICS",
    7: "THERMAL AND MOISTURE PROTECTION",
    8: "DOORS AND WINDOWS",
    9: "FINISHES",
    10: "SPECIALTIES",
    11: "EQUIPMENT",
    12: "FURNISHINGS",
    13: "SPECIAL CONSTRUCTION",
    14: "CONVEYING SYSTEMS",
    15: "MECHANICAL",
    16: "ELECTRICAL",
    17: "CHANGES",
}

KIND_LABEL = {
    "self": "Self-perform",
    "subtrade": "Subtrade",
    "quote": "Quote",
    "pc": "P.C. allowance",
    "excluded": "Excluded",
}

HEAD = Font(bold=True, color="FFFFFF")
HEAD_FILL = PatternFill("solid", fgColor="333333")
DIV_FILL = PatternFill("solid", fgColor="DDDDDD")
BOLD = Font(bold=True)
MONEY = '#,##0.00'
THIN = Border(bottom=Side(style="thin", color="BBBBBB"))


def division_of(code):
    """'9.12' -> 9. Codes are strings; never parse them as floats."""
    try:
        return int(str(code).split(".")[0])
    except (ValueError, AttributeError):
        return 99


def line_material(line):
    """Material/subtrade cost. Excluded lines contribute nothing."""
    if line.get("kind") == "excluded":
        return 0.0
    return float(line.get("qty", 0)) * float(line.get("rate", 0))


def line_labour(line, blended_rate):
    if line.get("kind") == "excluded":
        return 0.0
    return float(line.get("hours", 0)) * blended_rate


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_header_rows(ws, data, subtitle):
    ws["A1"] = data.get("prepared_by", "Caisley Developments Ltd.")
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = data.get("project", "")
    ws["A2"].font = Font(bold=True, size=12)
    ws["A3"] = subtitle
    ws["A4"] = f"Date: {data.get('date', '')}"
    ws["A5"] = f"Drawings: {data.get('drawings', 'not stated')}"
    if data.get("drawings_current") is False:
        ws["A6"] = ("NOT CONFIRMED AS THE CURRENT DRAWING SET — "
                    "priced as issued above")
        ws["A6"].font = Font(bold=True, color="AA0000")
    return 8


def sheet_internal(wb, data):
    ws = wb.create_sheet("ESTIMATE DETAILS")
    rate = float(data.get("blended_rate", 68.21))
    row = write_header_rows(ws, data, "INTERNAL WORKING ESTIMATE")

    headers = ["Code", "Description", "Qty", "Unit", "Rate",
               "Material/Sub", "Hours", "Labour", "Type", "Source"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row, col, h)
        c.font, c.fill = HEAD, HEAD_FILL
    row += 1

    lines = sorted(data.get("lines", []),
                   key=lambda l: (division_of(l.get("code")), str(l.get("code"))))
    totals = {}

    current_div = None
    for line in lines:
        div = division_of(line.get("code"))
        if div != current_div:
            current_div = div
            label = f"DIVISION {div} - {DIVISIONS.get(div, 'UNCLASSIFIED')}"
            c = ws.cell(row, 1, label)
            c.font = BOLD
            for col in range(1, len(headers) + 1):
                ws.cell(row, col).fill = DIV_FILL
            row += 1

        mat = line_material(line)
        lab = line_labour(line, rate)
        totals.setdefault(div, [0.0, 0.0])
        totals[div][0] += mat
        totals[div][1] += lab

        ws.cell(row, 1, line.get("code"))
        ws.cell(row, 2, line.get("desc"))
        if line.get("kind") == "excluded":
            # Cost columns carry the exclusion wording, as in the source workbook.
            ws.cell(row, 6, line.get("note", "NOT INCLUDED")).font = Font(italic=True)
        else:
            ws.cell(row, 3, line.get("qty"))
            ws.cell(row, 4, line.get("unit"))
            ws.cell(row, 5, line.get("rate")).number_format = MONEY
            ws.cell(row, 6, mat).number_format = MONEY
            if line.get("hours"):
                ws.cell(row, 7, line["hours"])
                ws.cell(row, 8, lab).number_format = MONEY
        ws.cell(row, 9, KIND_LABEL.get(line.get("kind"), line.get("kind", "")))
        ws.cell(row, 10, line.get("source", ""))
        for col in range(1, len(headers) + 1):
            ws.cell(row, col).border = THIN
        row += 1

        if line.get("vendor"):
            ws.cell(row, 2, f"    {line['vendor']}").font = Font(italic=True, size=9)
            row += 1

    autosize(ws, [9, 52, 11, 12, 12, 15, 9, 13, 15, 10])
    return totals


def sheet_summary(wb, data, totals):
    ws = wb.create_sheet("SUMMARY", 0)
    row = write_header_rows(ws, data, "CLIENT ESTIMATE")

    for col, h in enumerate(["", "Division", "Amount"], start=1):
        c = ws.cell(row, col, h)
        c.font, c.fill = HEAD, HEAD_FILL
    row += 1

    subtotal = 0.0
    for div in sorted(totals):
        amount = totals[div][0] + totals[div][1]
        subtotal += amount
        ws.cell(row, 2, f"DIVISION {div} - {DIVISIONS.get(div, 'UNCLASSIFIED')}")
        ws.cell(row, 3, amount).number_format = MONEY
        row += 1

    contingency_pct = float(data.get("contingency", 0.05))
    fee_pct = float(data.get("fee", 0.10))
    gst_pct = float(data.get("gst", 0.05))

    # Fee compounds on the contingency-inclusive subtotal; GST sits below the fee.
    contingency = subtotal * contingency_pct
    after_cont = subtotal + contingency
    fee = after_cont * fee_pct
    after_fee = after_cont + fee
    gst = after_fee * gst_pct
    total = after_fee + gst

    row += 1
    for label, value, bold in [
        ("SUB-TOTAL", subtotal, True),
        (f"CONTINGENCY / INFLATION @ {contingency_pct:.1%}", contingency, False),
        ("SUB-TOTAL INCLUDING CONTINGENCY", after_cont, True),
        (f"CONTRACTOR'S FEE @ {fee_pct:.1%}", fee, False),
        ("SUB-TOTAL INCLUDING FEE", after_fee, True),
        (f"GST @ {gst_pct:.1%}", gst, False),
        ("TOTAL ESTIMATE", total, True),
    ]:
        ws.cell(row, 2, label).font = BOLD if bold else Font()
        c = ws.cell(row, 3, value)
        c.number_format = MONEY
        if bold:
            c.font = BOLD
        row += 1

    autosize(ws, [3, 52, 18])
    return subtotal, total


def sheet_list(wb, title, heading, items, formatter=str):
    if not items:
        return
    ws = wb.create_sheet(title)
    ws["A1"] = heading
    ws["A1"].font = Font(bold=True, size=12)
    row = 3
    for item in items:
        ws.cell(row, 1, f"{row - 2}.")
        c = ws.cell(row, 2, formatter(item))
        c.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1
    autosize(ws, [5, 100])


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("takeoff", type=Path)
    ap.add_argument("-o", "--out", type=Path, default=Path("estimate.xlsx"))
    args = ap.parse_args()

    if not args.takeoff.exists():
        sys.exit(f"not found: {args.takeoff}")
    data = json.loads(args.takeoff.read_text())
    if not data.get("lines"):
        sys.exit("takeoff has no lines")

    def shared_lists(wb):
        sheet_list(wb, "ASSUMPTIONS", "Assumptions", data.get("assumptions", []))
        sheet_list(wb, "EXCLUSIONS", "Exclusions", data.get("exclusions", []))
        sheet_list(
            wb, "QUESTIONS", "Open questions", data.get("questions", []),
            lambda q: (f"{q.get('q', '')}\n"
                       f"Assumed meanwhile: {q.get('assumed', '—')}\n"
                       f"Affects: {q.get('affects', '—')}"),
        )

    # Two separate files. The detail must never travel in the same workbook as
    # the client summary — a second tab is one click from exposing every rate
    # and the margin.
    stem = args.out.with_suffix("")
    internal_path = Path(f"{stem}-internal.xlsx")
    client_path = Path(f"{stem}-client.xlsx")

    wb_int = Workbook()
    wb_int.remove(wb_int.active)
    totals = sheet_internal(wb_int, data)
    subtotal, total = sheet_summary(wb_int, data, totals)
    shared_lists(wb_int)
    # Change-order exposure stays internal — it is commercially sensitive.
    sheet_list(
        wb_int, "RISK", "Change-order exposure", data.get("risks", []),
        lambda r: f"{r.get('item', '')} — {r.get('why', '')}",
    )
    wb_int.save(internal_path)

    wb_cli = Workbook()
    wb_cli.remove(wb_cli.active)
    sheet_summary(wb_cli, data, totals)
    shared_lists(wb_cli)
    wb_cli.save(client_path)

    labour = sum(v[1] for v in totals.values())
    print(f"internal  {internal_path}")
    print(f"client    {client_path}")
    print()
    print(f"  materials + subtrades  {subtotal - labour:>14,.2f}")
    # Quoted against the total, matching the house convention (18.97% on 2033 E. 7).
    print(f"  contractor's labour    {labour:>14,.2f}  ({labour / total:.2%} of total estimate)")
    print(f"  trade subtotal         {subtotal:>14,.2f}")
    print(f"  TOTAL incl. fee + GST  {total:>14,.2f}")
    print(f"\nBoth files reconcile to {total:,.2f}. The client file holds division "
          f"totals only — no rates, hours or margin.")


if __name__ == "__main__":
    main()
